import string
import random
import openpyxl

def generate_password(length):
    characters = string.ascii_letters + string.digits + string.punctuation
    password = ''.join(random.choice(characters) for i in range(length))
    return password

name = input("The name of the account for the password: ")
password_length = int(input("Enter the length of the password: "))

password = generate_password(password_length)

try:
    workbook = openpyxl.load_workbook('passwords.xlsx')
    worksheet = workbook.active
    last_row = worksheet.max_row + 1
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    last_row = worksheet.max_row

worksheet.cell(row=last_row, column=1, value=name)
worksheet.cell(row=last_row, column=2, value=password)

workbook.save('passwords.xlsx')
